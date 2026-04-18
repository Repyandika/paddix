import csv
import io
import json

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
from pydantic import BaseModel
from typing import Optional

from database import get_db
from models import User
from auth_utils import require_admin

router = APIRouter(prefix="/api/admin", tags=["Admin"])


# ─── Schemas ──────────────────────────────────────────────────────────────────
class SawahCreate(BaseModel):
    kecamatan: str
    luas_ha: float
    id_sawah: Optional[float] = None
    status_data: Optional[str] = "valid"
    geojson_geometry: dict  # GeoJSON geometry object

class SawahUpdateGeometry(BaseModel):
    geojson_geometry: dict  # GeoJSON geometry object


# ──────────────────────────────────────────────────────────────────────────────
# HELPER: Recalculate KPI setelah CRUD sawah
# ──────────────────────────────────────────────────────────────────────────────
def _refresh_kpi(db: Session, kecamatan: str):
    """
    Recalculate kpi_kecamatan berdasarkan data aktual di sawah_karawang.
    Menggunakan UPDATE karena kolom fid NOT NULL (primary key).
    """
    db.execute(text("""
        UPDATE kpi_kecamatan
        SET
            "count" = sub.cnt,
            "sum"   = sub.s,
            "mean"  = sub.m,
            "median"= sub.med,
            "stddev"= sub.sd,
            "min"   = sub.mn,
            "max"   = sub.mx
        FROM (
            SELECT
                COUNT(*)::int            AS cnt,
                COALESCE(SUM(luas_ha),0) AS s,
                COALESCE(AVG(luas_ha),0) AS m,
                COALESCE(PERCENTILE_CONT(0.5) WITHIN GROUP (ORDER BY luas_ha),0) AS med,
                COALESCE(STDDEV(luas_ha),0)  AS sd,
                COALESCE(MIN(luas_ha),0)     AS mn,
                COALESCE(MAX(luas_ha),0)     AS mx
            FROM sawah_karawang
            WHERE LOWER(kecamatan) = LOWER(:kec)
        ) sub
        WHERE LOWER(kpi_kecamatan.kecamatan) = LOWER(:kec)
    """), {"kec": kecamatan})
    db.commit()


# ──────────────────────────────────────────────────────────────────────────────
# POST /api/admin/sawah  — Tambah poligon sawah baru
# ──────────────────────────────────────────────────────────────────────────────
@router.post("/sawah", summary="Tambah poligon sawah baru (admin only)")
def create_sawah(body: SawahCreate, admin: User = Depends(require_admin), db: Session = Depends(get_db)):
    import time
    
    # Auto-generate id_sawah (numeric, 6 digits max usually) if not provided
    generated_id_sawah = body.id_sawah
    if not generated_id_sawah:
        generated_id_sawah = float(int(time.time() * 100) % 899999 + 100000)

    geojson_str = json.dumps(body.geojson_geometry)
    sql = text("""
        INSERT INTO sawah_karawang (kecamatan, luas_ha, id_sawah, status_data, wkb_geometry)
        VALUES (
            :kecamatan, 
            COALESCE(ST_Area(ST_SetSRID(ST_GeomFromGeoJSON(:geojson), 4326)::geography) / 10000.0, 0),
            :id_sawah, 
            :status_data, 
            ST_Multi(ST_Force3D(ST_SetSRID(ST_GeomFromGeoJSON(:geojson), 4326)))
        )
        RETURNING ogc_fid
    """)
    result = db.execute(sql, {
        "kecamatan": body.kecamatan,
        "id_sawah": generated_id_sawah,
        "status_data": body.status_data,
        "geojson": geojson_str,
    })
    db.commit()
    new_id = result.fetchone()[0]

    # Auto-refresh KPI untuk kecamatan ini
    _refresh_kpi(db, body.kecamatan)

    return {"detail": "Poligon sawah berhasil ditambahkan", "ogc_fid": new_id}


# ──────────────────────────────────────────────────────────────────────────────
# PUT /api/admin/sawah/{ogc_fid}  — Edit geometri poligon
# ──────────────────────────────────────────────────────────────────────────────
@router.put("/sawah/{ogc_fid}", summary="Edit geometri poligon sawah (admin only)")
def update_sawah_geometry(ogc_fid: int, body: SawahUpdateGeometry, admin: User = Depends(require_admin), db: Session = Depends(get_db)):
    # Cek apakah poligon ada & ambil nama kecamatannya
    check = db.execute(text("SELECT ogc_fid, kecamatan FROM sawah_karawang WHERE ogc_fid = :fid"), {"fid": ogc_fid}).fetchone()
    if not check:
        raise HTTPException(status_code=404, detail=f"Poligon dengan ogc_fid={ogc_fid} tidak ditemukan")

    geojson_str = json.dumps(body.geojson_geometry)

    # Update geometri dan hitung ulang luas
    sql = text("""
        UPDATE sawah_karawang
        SET wkb_geometry = ST_Multi(ST_Force3D(ST_SetSRID(ST_GeomFromGeoJSON(:geojson), 4326))),
            luas_ha = COALESCE(ST_Area(ST_SetSRID(ST_GeomFromGeoJSON(:geojson), 4326)::geography) / 10000.0, 0)
        WHERE ogc_fid = :fid
    """)
    db.execute(sql, {"geojson": geojson_str, "fid": ogc_fid})
    db.commit()

    # Auto-refresh KPI untuk kecamatan ini
    _refresh_kpi(db, check.kecamatan)

    return {"detail": f"Geometri poligon ogc_fid={ogc_fid} berhasil diperbarui"}


# ──────────────────────────────────────────────────────────────────────────────
# DELETE /api/admin/sawah/{ogc_fid}  — Hapus poligon
# ──────────────────────────────────────────────────────────────────────────────
@router.delete("/sawah/{ogc_fid}", summary="Hapus poligon sawah (admin only)")
def delete_sawah(ogc_fid: int, admin: User = Depends(require_admin), db: Session = Depends(get_db)):
    # Ambil kecamatan sebelum dihapus (untuk refresh KPI)
    check = db.execute(text("SELECT ogc_fid, kecamatan FROM sawah_karawang WHERE ogc_fid = :fid"), {"fid": ogc_fid}).fetchone()
    if not check:
        raise HTTPException(status_code=404, detail=f"Poligon dengan ogc_fid={ogc_fid} tidak ditemukan")

    kecamatan_name = check.kecamatan

    db.execute(text("DELETE FROM sawah_karawang WHERE ogc_fid = :fid"), {"fid": ogc_fid})
    db.commit()

    # Auto-refresh KPI untuk kecamatan ini
    _refresh_kpi(db, kecamatan_name)

    return {"detail": f"Poligon ogc_fid={ogc_fid} berhasil dihapus"}



# ──────────────────────────────────────────────────────────────────────────────
# GET /api/admin/sawah/export-csv  — Export CSV
# ──────────────────────────────────────────────────────────────────────────────
@router.get("/sawah/export-csv", summary="Export data sawah ke CSV (admin only)")
def export_csv(admin: User = Depends(require_admin), db: Session = Depends(get_db)):
    sql = text("""
        SELECT ogc_fid, kecamatan, luas_ha, id_sawah, status_data,
               ST_AsText(wkb_geometry) as wkt_geometry
        FROM sawah_karawang
        ORDER BY kecamatan, ogc_fid
    """)
    rows = db.execute(sql).fetchall()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ogc_fid", "kecamatan", "luas_ha", "id_sawah", "status_data", "wkt_geometry"])
    for row in rows:
        writer.writerow([row.ogc_fid, row.kecamatan, row.luas_ha, row.id_sawah, row.status_data, row.wkt_geometry])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=sawah_karawang_export.csv"}
    )


# ──────────────────────────────────────────────────────────────────────────────
# POST /api/admin/ndvi/import-csv  — Import CSV Data NDVI
# ──────────────────────────────────────────────────────────────────────────────
@router.post("/ndvi/import-csv", summary="Import data NDVI bulanan dari CSV (admin only)")
async def import_ndvi_csv(file: UploadFile = File(...), admin: User = Depends(require_admin), db: Session = Depends(get_db)):
    """
    Format CSV yang diharapkan:
    kecamatan,periode,tahun,bulan,mean_ndvi,std_ndvi,pixel_count,jumlah_citra,kategori
    """
    if not file.filename.endswith(".csv"):
        raise HTTPException(status_code=400, detail="File harus berformat .csv")

    content = await file.read()
    decoded = content.decode("utf-8")
    reader = csv.DictReader(io.StringIO(decoded))

    updated = 0
    errors = []
    
    for i, row in enumerate(reader, start=2):
        try:
            kecamatan = row.get("kecamatan", "").strip()
            periode = row.get("periode", "").strip()
            if not kecamatan or not periode:
                errors.append(f"Baris {i}: 'kecamatan' atau 'periode' kosong")
                continue

            tahun = int(row.get("tahun", 0)) if row.get("tahun") else None
            bulan = int(row.get("bulan", 0)) if row.get("bulan") else None
            mean_ndvi = float(row.get("mean_ndvi", 0)) if row.get("mean_ndvi") else None
            std_ndvi = float(row.get("std_ndvi", 0)) if row.get("std_ndvi") else 0
            pixel_count = int(row.get("pixel_count", 0)) if row.get("pixel_count") else 0
            jumlah_citra = int(row.get("jumlah_citra", 0)) if row.get("jumlah_citra") else 0
            kategori = row.get("kategori", "").strip() or None

            # Upsert Data NDVI
            sql_upsert = text("""
                INSERT INTO ndvi_kecamatan 
                (kecamatan, periode, tahun, bulan, mean_ndvi, std_ndvi, pixel_count, jumlah_citra, kategori, created_at)
                VALUES 
                (:kecamatan, :periode, :tahun, :bulan, :mean_ndvi, :std_ndvi, :pixel_count, :jumlah_citra, :kategori, NOW())
                ON CONFLICT (kecamatan, periode) 
                DO UPDATE SET 
                    tahun = EXCLUDED.tahun,
                    bulan = EXCLUDED.bulan,
                    mean_ndvi = EXCLUDED.mean_ndvi,
                    std_ndvi = EXCLUDED.std_ndvi,
                    pixel_count = EXCLUDED.pixel_count,
                    jumlah_citra = EXCLUDED.jumlah_citra,
                    kategori = EXCLUDED.kategori,
                    created_at = NOW()
            """)
            db.execute(sql_upsert, {
                "kecamatan": kecamatan, "periode": periode, "tahun": tahun, "bulan": bulan,
                "mean_ndvi": mean_ndvi, "std_ndvi": std_ndvi, "pixel_count": pixel_count,
                "jumlah_citra": jumlah_citra, "kategori": kategori
            })
            
            updated += 1
        except Exception as e:
            errors.append(f"Baris {i} Gagal: {str(e)}")

    db.commit()
    return {
        "detail": f"Import data NDVI selesai. {updated} baris diproses.",
        "errors": errors[:20],
    }


# ──────────────────────────────────────────────────────────────────────────────
# GET /api/admin/ndvi/export-csv  — Export NDVI per kecamatan (ringan)
# ──────────────────────────────────────────────────────────────────────────────
@router.get("/ndvi/export-csv", summary="Export data NDVI per kecamatan ke CSV (admin only)")
def export_ndvi_csv(
    tahun: int = None,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    tahun_filter = "WHERE n.tahun = :tahun" if tahun else ""
    sql = text(f"""
        SELECT n.kecamatan, n.periode, n.tahun, n.bulan,
               n.mean_ndvi::float, n.std_ndvi::float,
               n.pixel_count, n.jumlah_citra, n.kategori
        FROM ndvi_kecamatan n
        {tahun_filter}
        ORDER BY n.kecamatan, n.tahun, n.bulan
    """)
    params = {"tahun": tahun} if tahun else {}
    rows = db.execute(sql, params).fetchall()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["kecamatan", "periode", "tahun", "bulan", "mean_ndvi", "std_ndvi", "pixel_count", "jumlah_citra", "kategori"])
    for row in rows:
        writer.writerow([row.kecamatan, row.periode, row.tahun, row.bulan, row.mean_ndvi, row.std_ndvi, row.pixel_count, row.jumlah_citra, row.kategori])

    output.seek(0)
    fname = f"ndvi_karawang_{tahun}.csv" if tahun else "ndvi_karawang_semua.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )


# ──────────────────────────────────────────────────────────────────────────────
# GET /api/admin/ndvi/export-filtered  — Export NDVI terfilter (tahun/bulan/kecamatan)
# ──────────────────────────────────────────────────────────────────────────────
@router.get("/ndvi/export-filtered", summary="Export NDVI terfilter ke CSV")
def export_ndvi_filtered(
    tahun: int = None,
    bulan: int = None,
    kecamatan: str = None,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    conditions = []
    params = {}
    if tahun:
        conditions.append("n.tahun = :tahun")
        params["tahun"] = tahun
    if bulan:
        conditions.append("n.bulan = :bulan")
        params["bulan"] = bulan
    if kecamatan:
        conditions.append("LOWER(n.kecamatan) = LOWER(:kecamatan)")
        params["kecamatan"] = kecamatan

    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""

    sql = text(f"""
        SELECT n.kecamatan, n.periode, n.mean_ndvi::float, n.kategori
        FROM ndvi_kecamatan n
        {where}
        ORDER BY n.kecamatan, n.tahun, n.bulan
    """)
    rows = db.execute(sql, params).fetchall()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["kecamatan", "periode", "mean_ndvi", "kategori"])
    for row in rows:
        writer.writerow([row.kecamatan, row.periode, row.mean_ndvi, row.kategori])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=ndvi_filtered.csv"}
    )


# ──────────────────────────────────────────────────────────────────────────────
# GET /api/admin/ranking/export-csv  — Export Ranking Risiko ke CSV
# ──────────────────────────────────────────────────────────────────────────────
@router.get("/ranking/export-csv", summary="Export ranking risiko kecamatan ke CSV")
def export_ranking_csv(
    tahun: int = None,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    tahun_filter = "WHERE n.tahun = :tahun" if tahun else ""
    params = {"tahun": tahun} if tahun else {}

    sql = text(f"""
        SELECT
            n.kecamatan,
            AVG(n.mean_ndvi)::float AS avg_ndvi
        FROM ndvi_kecamatan n
        {tahun_filter}
        GROUP BY n.kecamatan
        ORDER BY avg_ndvi ASC
    """)
    rows = db.execute(sql, params).fetchall()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Rank", "Kecamatan", "Mean NDVI", "Kategori"])
    for i, row in enumerate(rows, 1):
        v = row.avg_ndvi
        if v is None:
            kat = "Data Tidak Tersedia"
        elif v < 0.25:
            kat = "Risiko Tinggi"
        elif v < 0.40:
            kat = "Risiko Sedang"
        else:
            kat = "Normal / Aman"
        writer.writerow([i, row.kecamatan, round(v, 4) if v else None, kat])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=ranking_risiko.csv"}
    )


# ──────────────────────────────────────────────────────────────────────────────
# GET /api/admin/report/export-xlsx  — Export Laporan Analitik Excel
# ──────────────────────────────────────────────────────────────────────────────
@router.get("/report/export-xlsx", summary="Export laporan analitik lengkap ke Excel (.xlsx)")
def export_report_xlsx(
    tahun: int = None,
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    tahun_filter = "WHERE n.tahun = :tahun" if tahun else ""
    params = {"tahun": tahun} if tahun else {}

    # ── Sheet 1: Ranking Risiko ──
    ws1 = wb.active
    ws1.title = "Ranking Risiko"
    headers1 = ["Rank", "Kecamatan", "Mean NDVI", "Kategori"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="166534")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    sql_rank = text(f"""
        SELECT kecamatan, AVG(mean_ndvi)::float AS avg_ndvi
        FROM ndvi_kecamatan n {tahun_filter}
        GROUP BY kecamatan ORDER BY avg_ndvi ASC
    """)
    rows_rank = db.execute(sql_rank, params).fetchall()
    for i, row in enumerate(rows_rank, 1):
        v = row.avg_ndvi
        kat = "Data Tidak Tersedia" if v is None else ("Risiko Tinggi" if v < 0.25 else ("Risiko Sedang" if v < 0.40 else "Normal / Aman"))
        ws1.cell(row=i+1, column=1, value=i).border = thin_border
        ws1.cell(row=i+1, column=2, value=row.kecamatan).border = thin_border
        c = ws1.cell(row=i+1, column=3, value=round(v, 4) if v else None)
        c.border = thin_border
        c.number_format = "0.0000"
        kat_cell = ws1.cell(row=i+1, column=4, value=kat)
        kat_cell.border = thin_border
        if kat == "Risiko Tinggi":
            kat_cell.fill = PatternFill("solid", fgColor="FEE2E2")
            kat_cell.font = Font(color="DC2626", bold=True)
        elif kat == "Risiko Sedang":
            kat_cell.fill = PatternFill("solid", fgColor="FEF3C7")

    ws1.column_dimensions["A"].width = 8
    ws1.column_dimensions["B"].width = 22
    ws1.column_dimensions["C"].width = 14
    ws1.column_dimensions["D"].width = 20

    # ── Sheet 2: Data NDVI Bulanan ──
    ws2 = wb.create_sheet(title="Data NDVI Bulanan")
    headers2 = ["Kecamatan", "Periode", "Tahun", "Bulan", "Mean NDVI", "Std NDVI", "Pixel Count", "Kategori"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    sql_data = text(f"""
        SELECT kecamatan, periode, tahun, bulan, mean_ndvi::float, std_ndvi::float, pixel_count, kategori
        FROM ndvi_kecamatan n {tahun_filter}
        ORDER BY kecamatan, tahun, bulan
    """)
    rows_data = db.execute(sql_data, params).fetchall()
    for i, row in enumerate(rows_data, 2):
        ws2.cell(row=i, column=1, value=row.kecamatan)
        ws2.cell(row=i, column=2, value=row.periode)
        ws2.cell(row=i, column=3, value=row.tahun)
        ws2.cell(row=i, column=4, value=row.bulan)
        ws2.cell(row=i, column=5, value=row.mean_ndvi)
        ws2.cell(row=i, column=6, value=row.std_ndvi)
        ws2.cell(row=i, column=7, value=row.pixel_count)
        ws2.cell(row=i, column=8, value=row.kategori)

    # ── Sheet 3: Statistik Ringkasan ──
    ws3 = wb.create_sheet(title="Statistik Ringkasan")
    headers3 = ["Kecamatan", "Rata-Rata NDVI", "Min NDVI", "Max NDVI", "Std Dev", "Total Record"]
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    sql_stat = text(f"""
        SELECT kecamatan, AVG(mean_ndvi)::float, MIN(mean_ndvi)::float,
               MAX(mean_ndvi)::float, STDDEV(mean_ndvi)::float, COUNT(*)::int
        FROM ndvi_kecamatan n {tahun_filter}
        GROUP BY kecamatan ORDER BY kecamatan
    """)
    rows_stat = db.execute(sql_stat, params).fetchall()
    for i, row in enumerate(rows_stat, 2):
        ws3.cell(row=i, column=1, value=row[0])
        for j in range(1, 6):
            c = ws3.cell(row=i, column=j+1, value=round(row[j], 4) if row[j] else None)
            if j < 5:
                c.number_format = "0.0000"

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    fname = f"laporan_paddix_{tahun}.xlsx" if tahun else "laporan_paddix_semua.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )


